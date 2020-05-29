from openpyxl import *
from PyQt5.QtWidgets import QMessageBox, QFileDialog


# классы, который определяют шаблон загруженной таблицы и действия над ней
class Table:
    def __init__(self, path):
        # открыть файл
        self.workbook = self.load_table(path)
        if self.workbook is None:
            raise Exception("Нет таблицы")
        self.table_path = path

        self.worksheet = self.workbook.active # лист, с которым работаем

        # распознать шаблон таблицы
        self.pattern = TablePattern(self.worksheet[1])


    # открывает файл таблицы
    def load_table(self, path, message="Не удалось открыть файл. Выберите расположение таблицы: "):
        try:
            wb = load_workbook(path)
            return wb
        except:
            QMessageBox.critical(None, "Ошибка", message, QMessageBox.Open)
            path = QFileDialog.getOpenFileName(None, "Выберите расположение шаблона...", ".", "Exel (*.xlsx)")[0]
            if path == '':
                return None
            return self.load_table(path)

    # возвращает все колонки с их индексами
    def get_columns(self):
        return self.pattern.columns

    # возвращает следующую пустую строку после тех, что уже есть
    def get_next_empty_row(self):
        return self.worksheet.max_row + 1

    # вставка в ячейку
    def insert_into_cell(self, column, row, value):
        self.worksheet.cell(row, column).value = value

    # сохранить обновленную таблицу
    def save(self):
        self.workbook.save(self.table_path)
        self.workbook = Workbook(self.table_path)
        self.worksheet = self.workbook.active


# считывает столбцы таблицы и преобразует их координаты для работы
class TablePattern:
    def __init__(self, column_names_row):
        # задают связь между названиями колонок в таблице и их отображениями в программе
        orig_column_names = {
            'Откуда пришли': 'where from short',
            'viber group': 'viber group',
            '№ зак.': 'order num',
            'тел.': 'telephone',
            'Имя': 'name',
            'Дата оплаты': 'pay date',
            'Покупка': 'order',
            'Розница поставщика': 'provider price',
            'Скидка поставщика, %': 'discount',
            'Розница ZelenKvit': 'sell price',
            'Прибыль': 'income',
            'Адрес': 'adress',
            'Доп. контактные данные': 'more contacts',
            'Примечания': 'notes'
        }
        # словарь, который связывает название колонки в программе и ее индекс в таблице
        self.columns = {}

        for cell in column_names_row:
            if cell.value not in orig_column_names.keys():
                message = "Column name not identified: " + cell.value
                raise Exception(message)
            self.columns[orig_column_names[cell.value]] = cell.column
